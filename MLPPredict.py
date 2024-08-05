# -*- coding: utf-8 -*-
"""
Created on Mon Jan 22 11:51:05 2024

@author: Ben
"""
import torch
import numpy as np
import  os
import random
import matplotlib.pyplot as plt
from torch.utils.data import Dataset
from torch.utils.data import DataLoader
from torchvision import datasets
from torchvision.transforms import ToTensor
import matplotlib.pyplot as plt
from torch import nn
import scipy
from torchvision import datasets,transforms
import torchvision.models as models
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import csv
# File dir directs to the folder where all the files are stored
# File paths +directory is then found through accessing filePathNumbers
#  \ben\data + \file1 .....
# the file that the path directs to is csv file 3 entries : time,resfreq,labels
class ABMDataset(Dataset):
    def __init__(self,r_dir,filePathNumbers):
        self.root = r_dir
        self.filePathN = filePathNumbers#pd.read_csv(filePathNumbers, engine='python',on_bad_lines='warn', encoding='unicode_escape')
        
    def __len__(self):
        #num of file paths gives number of dataset
        x=load_workbook(self.filePathN)
        x = x.active
        col = x['A']
        lst =[cell.value for cell in col[1:]]
        return len(lst)
    
    def __getitem__(self,idx):
        #print(idx)
        workbook = load_workbook(filename=self.filePathN)
        sheet = workbook.active
        #fix it to this next 
        path = os.path.join(self.root)
        #path = os.path.join(self.root, str(sheet.cell(row=idx+1, column=1).value))
        res =os.path.join(path,"Moleculetest"+str(idx)+".txt")
        time =os.path.join(path,"Timetest"+str(idx)+".txt")
        var =os.path.join(path,"variablestest"+str(idx)+".txt")
        fileR = open(res, 'r+')
        fileT = open(time, 'r+')
        fileV = open(var, 'r+')
        resFreqData = torch.tensor([float(word) for line in fileR if line.strip() for word in line.split()])
        timeData = torch.tensor([float(word) for line in fileT if line.strip() for word in line.split()])
        variables = torch.tensor([float(word) for line in fileV if line.strip() for word in line.split()])
        #print("heh",variables.size())
        return resFreqData, timeData, variables
        # Test size
        # self.net = nn.Sequential(
        #     #nn.Flatten(),
        #     nn.Linear(6,32),
        #     nn.LeakyReLU(),
        #     nn.Linear(32,128),
        #     nn.LeakyReLU(),
        #     nn.Linear(128,256),
        #     nn.LeakyReLU(),
        #     nn.Linear(256,512),
        #     nn.LeakyReLU(),
        #     nn.Linear(512,256),
        #     nn.LeakyReLU(),
        #     nn.Linear(256,128),
        #     nn.LeakyReLU(),
        #     nn.Linear(128,64),
        #     nn.LeakyReLU(),
        #     nn.Linear(64,31))          
class NN(nn.Module):
    def __init__(self):
        super().__init__()
        self.net = nn.Sequential(
            #nn.Flatten(),
            nn.Linear(6,32),
            nn.LeakyReLU(),
            nn.Linear(32,128),
            nn.LeakyReLU(),
            nn.Linear(128,256),
            nn.LeakyReLU(),
            nn.Linear(256,512),
            nn.LeakyReLU(),
            nn.Linear(512,256),
            nn.LeakyReLU(),
            nn.Linear(256,128),
            nn.LeakyReLU(),
            nn.Linear(128,64),
            nn.LeakyReLU(),
            nn.Linear(64,31))
        
    def forward(self,x):
          logits = self.net(x)
          #logits = self.convNet(x)
       
          #print("A is",logits[0],"B is",logits[1],"C is ",logits[2])
          return logits

class LTSM(nn.Module):
    def __init__(self,inputDim,hiddenDim,outputDim,layerNum):
        super().__init__()
        self.inputSize =inputDim
        self.hiddenSize = hiddenDim
        self.outputDim = outputDim
        self.layerNum = layerNum
        self.LT =nn.LSTM(self.inputSize,self.hiddenSize,self.layerNum)
        self.LT2 =nn.LSTM(self.hiddenSize,self.hiddenSize,self.layerNum)
        self.lin = nn.Linear(self.inputSize,self.inputSize)
        self.lin2 = nn.Linear(self.hiddenSize,self.outputDim)

    def forward(self,x):
        #print(x.size())
        x1 = self.lin(x)
        #print("hello",x1.size())
        x2 ,x21 =self.LT(x1)
        #print(x2.size())
        x3,x31 =self.LT2(x2)
        logits =self.lin2(x3)

        return logits
     
def trainNN(dataset,model,optim,lossF):
    model.train()
    #Iterate through dataset
    for i, j in enumerate(dataset):
        #print(i)
        res,time,var = j
        optim.zero_grad()
        #print(var.size())
        #print(var.shape)
        pred = model(var)
        #print(pred.size())
        loss =lossF(pred,res)
        loss.backward()
        # Loss function difference between input and output
        optim.step()
        if(i == 1):
            return time
    
    return time

#model = NN()
model = LTSM(6,1024,31,10)
# Hyper parameters #####
learningRate = 1e-5
epochs =int(1)                   
rDir =r"C:\Users\Ben\OneDrive\Desktop\ABMDataset"
pathN = r"C:\Users\Ben\OneDrive\Desktop\ABMDataset\paths.xlsx"
lossF = nn.MSELoss()
optimiser = torch.optim.Adam(model.parameters(),lr =learningRate) 
# checkpoint = torch.load('LTSMDataFitmodel77.pth')
# model.load_state_dict(checkpoint['NN_state_dict'])
# optimiser.load_state_dict(checkpoint['NN_op_state_dict'])
data = ABMDataset(rDir,pathN)
dataset = DataLoader(data,drop_last = True)
maxBindingSite = 651
C0 = 115
n = 0.010219
Temp = 100
Kon = 1.1e2
Koff = 1.8e-2
for i in range(epochs):
    print(i)
    timeData =trainNN(dataset,model,optimiser,lossF)
x =torch.tensor([maxBindingSite,C0,n,Temp,Kon,Koff])
predData = model(x)
print(predData)
partData =pd.DataFrame(predData.detach().numpy())
partData.to_csv('partdata37.csv', index=False)
partData =pd.DataFrame(timeData.detach().numpy())
partData.to_csv('time37.csv', index=False)
torch.save({
            'NN_state_dict': model.state_dict(),
            'NN_op_state_dict': optimiser.state_dict(),
            },'LTSMDataFitmodel77.pth')
plt.figure(figsize = (10,8))
plt.title("The prediction of response using a NN")
plt.plot(timeData[0],predData.detach().numpy(),'b.',label = "Experiment Data")
plt.ylabel("Number of Analyte Bound")
plt.xlabel("Time (seconds) ")
plt.legend()
plt.show()